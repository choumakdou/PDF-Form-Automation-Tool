#!/usr/bin/env python3
"""
PDF Form Automation Tool  v1.0
==============================
通用 PDF 表格自動填寫工具

功能：
  1. 掃描任何 PDF 的互動欄位（文字、Checkbox、下拉等）
  2. 自動生成對應 Excel 輸入範本
  3. 讀取已填妥的 Excel，批量生成獨立 PDF（每行一份）

依賴：pip install pymupdf openpyxl
打包：pyinstaller --onefile --windowed pdf_form_tool.py
"""

import os, sys, re, threading
from pathlib import Path
from collections import defaultdict
from datetime import datetime

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

try:
    import fitz
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pymupdf"])
    import fitz

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════
#  資料結構
# ══════════════════════════════════════════════════════════

class FieldInfo:
    """代表 PDF 中一個邏輯欄位"""
    def __init__(self, col_key, display_name, field_type,
                 pdf_name, group_size=1, on_state="Yes",
                 options=None, hint=""):
        self.col_key      = col_key        # Excel 欄位鍵（唯一）
        self.display_name = display_name   # 顯示名稱（人類閱讀）
        self.field_type   = field_type     # Text / CheckBox / CheckBox_Group / Combo
        self.pdf_name     = pdf_name       # PDF 欄位原始名稱
        self.group_size   = group_size     # CheckBox_Group 的數量
        self.on_state     = on_state       # Checkbox 勾選值
        self.options      = options or []  # Combo 選項
        self.hint         = hint           # Excel 提示文字


# ══════════════════════════════════════════════════════════
#  PDF 分析
# ══════════════════════════════════════════════════════════

def _make_col_key(name: str, existing: set) -> str:
    """從 PDF 欄位名稱生成唯一的 Excel col_key"""
    # 取前 40 字元，移除特殊字元
    key = re.sub(r'[^\w\u4e00-\u9fff]', '_', name[:40]).strip('_')
    key = re.sub(r'_+', '_', key) or "field"
    base = key
    i = 2
    while key in existing:
        key = f"{base}_{i}"
        i += 1
    existing.add(key)
    return key


def analyze_pdf(pdf_path: str) -> list:
    """
    掃描 PDF，返回 FieldInfo 列表。
    自動過濾 Button（Print/Reset），
    合併同名 CheckBox 為 CheckBox_Group。
    重要：所有 widget 資料在 doc.close() 前讀取完畢。
    """
    doc = fitz.open(pdf_path)

    # 收集每個欄位的原始資料（在 doc 開啟時一次讀完）
    # raw: { pdf_name: { "ftype": str, "count": int, "on_state": str, "opts": list } }
    raw = {}
    order = []   # 保留欄位出現次序

    for page in doc:
        for w in page.widgets():
            name  = w.field_name
            ftype = w.field_type_string
            if ftype == "Button":
                continue
            if name not in raw:
                raw[name] = {"ftype": ftype, "count": 0,
                             "on_state": "Yes", "opts": []}
                order.append(name)
            raw[name]["count"] += 1

            # 讀取 on_state（只需第一個 widget 即可）
            if raw[name]["count"] == 1 and ftype == "CheckBox":
                try:
                    ov = w.on_state()
                    raw[name]["on_state"] = str(ov) if ov is not None else "Yes"
                except Exception:
                    raw[name]["on_state"] = "Yes"

            # 讀取 Combo 選項
            if ftype in ("Combo", "ListBox") and not raw[name]["opts"]:
                raw[name]["opts"] = w.choice_values or []

    doc.close()   # 安全關閉，不再存取 widget 物件

    # 根據收集到的原始資料建立 FieldInfo
    fields    = []
    used_keys = set()

    for pdf_name in order:
        info  = raw[pdf_name]
        ftype = info["ftype"]
        count = info["count"]

        col_key      = _make_col_key(pdf_name, used_keys)
        display_name = pdf_name if len(pdf_name) <= 45 else pdf_name[:42] + "..."

        if ftype == "Text":
            fi = FieldInfo(
                col_key=col_key,
                display_name=display_name,
                field_type="Text",
                pdf_name=pdf_name,
                hint="文字輸入"
            )

        elif ftype == "CheckBox":
            on_val = info["on_state"]
            if count == 1:
                fi = FieldInfo(
                    col_key=col_key,
                    display_name=display_name,
                    field_type="CheckBox",
                    pdf_name=pdf_name,
                    on_state=on_val,
                    hint="勾選填 Y；不勾留空"
                )
            else:
                fi = FieldInfo(
                    col_key=col_key,
                    display_name=display_name,
                    field_type="CheckBox_Group",
                    pdf_name=pdf_name,
                    group_size=count,
                    on_state=on_val,
                    hint=f"填序號 0–{count-1}（共 {count} 個選項）"
                )

        elif ftype in ("Combo", "ListBox"):
            opts = info["opts"]
            fi = FieldInfo(
                col_key=col_key,
                display_name=display_name,
                field_type="Combo",
                pdf_name=pdf_name,
                options=opts,
                hint="可選: " + " / ".join(str(o) for o in opts[:6])
            )

        else:
            fi = FieldInfo(
                col_key=col_key,
                display_name=display_name,
                field_type="Text",
                pdf_name=pdf_name,
                hint=f"({ftype})"
            )

        fields.append(fi)

    return fields


# ══════════════════════════════════════════════════════════
#  Excel 範本生成
# ══════════════════════════════════════════════════════════

def generate_excel_template(fields: list, pdf_path: str, out_path: str):
    """
    生成 Excel 輸入範本。
    Row 1 : 標題
    Row 2 : 填寫說明
    Row 3 : col_key（機器讀取用，淺灰底）
    Row 4 : 欄位說明 / 提示（人類閱讀，黃底）
    Row 5+ : 資料輸入區
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # ── 顏色 / 樣式 ──
    C_DARK   = "1F4E79"
    C_MED    = "2E75B6"
    C_GREEN  = "375623"
    C_KEY    = "D6E4F0"
    C_HINT   = "FFF2CC"
    C_WHITE  = "FFFFFF"
    C_GRAY   = "F2F2F2"
    thin     = Side(border_style="thin", color="B8CCE4")
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)

    pdf_name   = Path(pdf_path).name
    total_cols = len(fields) + 1  # +1 for __output_name__

    # ── Row 1: 標題 ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t = ws.cell(1, 1)
    t.value     = f"PDF Form Automation Tool  ·  {pdf_name}  ·  每行生成一份 PDF"
    t.font      = Font(bold=True, color=C_WHITE, size=11, name="Calibri")
    t.fill      = PatternFill("solid", fgColor=C_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # ── Row 2: 說明 ──
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    n = ws.cell(2, 1)
    n.value = ("【填寫說明】每行代表一份 PDF。"
               "第 1 欄【輸出檔名】可留空（自動命名 Row_001.pdf）。"
               "CheckBox 欄填 Y 表示勾選；CheckBox_Group 欄填序號（0 起計）。")
    n.font      = Font(italic=True, color="7F6000", size=8, name="Calibri")
    n.fill      = PatternFill("solid", fgColor=C_HINT)
    n.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 26

    # ── 欄位定義（含固定首欄 __output_name__）──
    col_defs = [("__output_name__", "輸出檔名\n(不含 .pdf)", "留空則自動命名", "Meta")]
    for fi in fields:
        col_defs.append((fi.col_key, fi.display_name, fi.hint, fi.field_type))

    # ── Row 3: col_key ──
    for c, (key, disp, hint, ftype) in enumerate(col_defs, 1):
        cell            = ws.cell(3, c)
        cell.value      = key
        cell.font       = Font(bold=True, color=C_DARK, size=7, name="Consolas")
        cell.fill       = PatternFill("solid", fgColor=C_KEY)
        cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border     = bdr

    # ── Row 4: 顯示名稱（人類閱讀）──
    for c, (key, disp, hint, ftype) in enumerate(col_defs, 1):
        cell            = ws.cell(4, c)
        cell.value      = disp
        bg = C_GREEN if "CheckBox" in ftype else C_MED
        bg = C_DARK  if ftype == "Meta"     else bg
        cell.font       = Font(bold=True, color=C_WHITE, size=8, name="Calibri")
        cell.fill       = PatternFill("solid", fgColor=bg)
        cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border     = bdr

    # ── Row 5: 提示行 ──
    for c, (key, disp, hint, ftype) in enumerate(col_defs, 1):
        cell            = ws.cell(5, c)
        cell.value      = hint
        cell.font       = Font(italic=True, color="595959", size=7, name="Calibri")
        cell.fill       = PatternFill("solid", fgColor=C_GRAY)
        cell.alignment  = Alignment(wrap_text=True, vertical="top")
        cell.border     = bdr

    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 38
    ws.row_dimensions[5].height = 28

    # ── Row 6–25: 空白資料行 ──
    for row in range(6, 26):
        for col in range(1, len(col_defs) + 1):
            c           = ws.cell(row, col)
            c.border    = bdr
            c.alignment = Alignment(vertical="top", wrap_text=False)
        ws.row_dimensions[row].height = 18

    # ── 欄寬 ──
    ws.column_dimensions["A"].width = 22
    for c in range(2, len(col_defs) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20

    ws.freeze_panes = "B6"

    # ── Field Guide 工作表 ──
    ws2 = wb.create_sheet("Field Guide")
    gh  = ["col_key（Excel欄）", "PDF 欄位全名", "類型", "提示 / 可選值"]
    for c, h in enumerate(gh, 1):
        cell       = ws2.cell(1, c, h)
        cell.font  = Font(bold=True, color=C_WHITE)
        cell.fill  = PatternFill("solid", fgColor=C_DARK)
        cell.alignment = Alignment(wrap_text=True)

    ws2.cell(2, 1, "__output_name__")
    ws2.cell(2, 2, "(特殊欄位)")
    ws2.cell(2, 3, "Meta")
    ws2.cell(2, 4, "輸出 PDF 檔名，不含副檔名。留空自動命名。")

    for r, fi in enumerate(fields, 3):
        ws2.cell(r, 1, fi.col_key)
        ws2.cell(r, 2, fi.pdf_name)
        ws2.cell(r, 3, fi.field_type)
        ws2.cell(r, 4, fi.hint)

    for col_letter, w in zip("ABCD", [28, 60, 18, 55]):
        ws2.column_dimensions[col_letter].width = w

    wb.save(out_path)


# ══════════════════════════════════════════════════════════
#  Excel 讀取
# ══════════════════════════════════════════════════════════

def read_excel_data(excel_path: str) -> tuple:
    """
    讀取資料工作表。
    Row 3 = col_keys，Row 6+ = 資料。
    返回 (col_keys: list, rows: list[dict])
    """
    wb  = openpyxl.load_workbook(excel_path, data_only=True)
    ws  = wb["Data"] if "Data" in wb.sheetnames else wb.active

    # 從 Row 3 讀取 col_keys
    col_keys = []
    for cell in ws[3]:
        col_keys.append(cell.value or "")

    # 從 Row 6 開始讀取資料
    rows = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        record = {}
        for idx, val in enumerate(row):
            if idx < len(col_keys) and col_keys[idx]:
                record[col_keys[idx]] = val
        rows.append(record)

    return col_keys, rows


# ══════════════════════════════════════════════════════════
#  PDF 填寫
# ══════════════════════════════════════════════════════════

def _safe_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def fill_pdf(template_path: str, fields: list, row_data: dict, out_path: str):
    """
    用 row_data 填寫 PDF 並儲存至 out_path。
    策略：迭代頁面 widget 時即時填寫，不預先緩存 widget 物件，
    避免 update() 後出現 'Annot is not bound to a page' 錯誤。
    """
    # 預先建立查找表：pdf_name → (FieldInfo, value_str)
    fill_map: dict = {}
    for fi in fields:
        val_str = _safe_str(row_data.get(fi.col_key))
        if val_str:
            fill_map[fi.pdf_name] = (fi, val_str)

    if not fill_map:
        # 無需填寫，直接複製
        import shutil
        shutil.copy2(template_path, out_path)
        return

    doc = fitz.open(template_path)

    # 記錄每個 pdf_name 已遇到的次數（用於 CheckBox_Group 的索引）
    name_counter = defaultdict(int)

    for page in doc:
        # 每次從頁面重新取得 widget 列表，避免 stale reference
        for widget in page.widgets():
            ftype = widget.field_type_string
            if ftype == "Button":
                continue

            fname = widget.field_name
            idx   = name_counter[fname]   # 本 widget 在同名欄位中的序號
            name_counter[fname] += 1

            if fname not in fill_map:
                continue

            fi, val_str = fill_map[fname]

            if fi.field_type == "Text":
                if idx == 0:              # 同名 Text 只填第一個
                    widget.field_value = val_str
                    widget.update()

            elif fi.field_type == "CheckBox":
                if idx == 0 and val_str.lower() in (
                        "y", "yes", "✓", "true", "1", "x", "是"):
                    widget.field_value = fi.on_state
                    widget.update()

            elif fi.field_type == "CheckBox_Group":
                try:
                    target = int(float(val_str))
                    if idx == target:
                        widget.field_value = fi.on_state
                        widget.update()
                except (ValueError, TypeError):
                    pass

            elif fi.field_type in ("Combo", "ListBox"):
                if idx == 0:
                    widget.field_value = val_str
                    widget.update()

    doc.save(out_path, garbage=3, deflate=True)
    doc.close()


def batch_fill_pdfs(template_path: str, fields: list, rows: list,
                    out_dir: str, progress_cb=None) -> list:
    """
    批量生成 PDF。
    progress_cb(done, total, filename, ok, err_msg) 用於更新 UI。
    返回 [(filename, ok, msg), ...]
    """
    results = []
    total   = len(rows)
    os.makedirs(out_dir, exist_ok=True)

    for i, row in enumerate(rows, 1):
        name_val  = _safe_str(row.get("__output_name__"))
        filename  = f"{name_val}.pdf" if name_val else f"Row_{i:03d}.pdf"
        out_path  = os.path.join(out_dir, filename)
        try:
            fill_pdf(template_path, fields, row, out_path)
            results.append((filename, True, ""))
            if progress_cb:
                progress_cb(i, total, filename, True, "")
        except Exception as e:
            results.append((filename, False, str(e)))
            if progress_cb:
                progress_cb(i, total, filename, False, str(e))

    return results


# ══════════════════════════════════════════════════════════
#  GUI
# ══════════════════════════════════════════════════════════

class App(tk.Tk):
    DARK  = "#1F4E79"
    MED   = "#2E75B6"
    LIGHT = "#EBF3FB"
    WHITE = "#FFFFFF"
    GREEN = "#375623"
    RED   = "#C00000"
    GRAY  = "#F5F5F5"

    def __init__(self):
        super().__init__()
        self.title("PDF Form Automation Tool  v1.0")
        self.resizable(True, True)
        self.minsize(680, 560)
        self.configure(bg=self.DARK)

        # 狀態變數
        self.pdf_path    = tk.StringVar()
        self.excel_path  = tk.StringVar()
        self.out_dir     = tk.StringVar()
        self.fields      = []   # list[FieldInfo]
        self.rows        = []   # list[dict]

        self._build_ui()
        self._center_window()

    # ── 介面建構 ──────────────────────────────────────────

    def _build_ui(self):
        # 頂部標題列
        header = tk.Frame(self, bg=self.DARK, pady=10)
        header.pack(fill="x")
        tk.Label(header, text="PDF Form Automation Tool",
                 font=("Calibri", 16, "bold"),
                 bg=self.DARK, fg=self.WHITE).pack(side="left", padx=18)
        tk.Label(header, text="v1.0",
                 font=("Calibri", 10),
                 bg=self.DARK, fg="#9DC3E6").pack(side="left")

        # 主體
        body = tk.Frame(self, bg=self.LIGHT, padx=18, pady=14)
        body.pack(fill="both", expand=True)

        # ── 步驟 1：PDF 範本 ──
        self._step_frame(body, "① 選擇 PDF 範本", 0)
        f1 = self._step_body(body, 0)
        tk.Entry(f1, textvariable=self.pdf_path, width=52,
                 font=("Calibri", 9)).pack(side="left", padx=(0, 6))
        tk.Button(f1, text="瀏覽…", command=self._browse_pdf,
                  **self._btn_style()).pack(side="left", padx=(0, 8))
        tk.Button(f1, text="掃描欄位 →", command=self._scan_fields,
                  **self._btn_style(primary=True)).pack(side="left")

        self.lbl_scan = tk.Label(body, text="", font=("Calibri", 9),
                                 bg=self.LIGHT, fg=self.GREEN)
        self.lbl_scan.pack(anchor="w", padx=8, pady=(0, 6))

        # ── 步驟 2：匯出範本 ──
        self._step_frame(body, "② 匯出 Excel 輸入範本", 1)
        f2 = self._step_body(body, 1)
        tk.Button(f2, text="匯出 Excel 範本…", command=self._export_template,
                  **self._btn_style(primary=True)).pack(side="left")
        self.lbl_export = tk.Label(f2, text="", font=("Calibri", 9),
                                   bg=self.LIGHT, fg=self.GREEN)
        self.lbl_export.pack(side="left", padx=10)

        # ── 步驟 3：載入已填 Excel ──
        self._step_frame(body, "③ 載入已填妥的 Excel", 2)
        f3 = self._step_body(body, 2)
        tk.Entry(f3, textvariable=self.excel_path, width=52,
                 font=("Calibri", 9)).pack(side="left", padx=(0, 6))
        tk.Button(f3, text="瀏覽…", command=self._browse_excel,
                  **self._btn_style()).pack(side="left", padx=(0, 8))
        tk.Button(f3, text="載入資料 →", command=self._load_excel,
                  **self._btn_style(primary=True)).pack(side="left")

        self.lbl_load = tk.Label(body, text="", font=("Calibri", 9),
                                 bg=self.LIGHT, fg=self.GREEN)
        self.lbl_load.pack(anchor="w", padx=8, pady=(0, 6))

        # ── 步驟 4：輸出資料夾 + 生成 ──
        self._step_frame(body, "④ 選擇輸出資料夾並生成 PDF", 3)
        f4 = self._step_body(body, 3)
        tk.Entry(f4, textvariable=self.out_dir, width=52,
                 font=("Calibri", 9)).pack(side="left", padx=(0, 6))
        tk.Button(f4, text="瀏覽…", command=self._browse_out,
                  **self._btn_style()).pack(side="left", padx=(0, 8))

        f4b = tk.Frame(body, bg=self.LIGHT)
        f4b.pack(fill="x", padx=8, pady=6)
        self.btn_gen = tk.Button(f4b, text="▶  生成所有 PDF",
                                 command=self._start_generate,
                                 font=("Calibri", 11, "bold"),
                                 bg=self.DARK, fg=self.WHITE,
                                 activebackground=self.MED,
                                 relief="flat", padx=18, pady=6,
                                 cursor="hand2")
        self.btn_gen.pack(side="left")

        # 進度列
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(f4b, variable=self.progress_var,
                                            maximum=100, length=260)
        self.progress_bar.pack(side="left", padx=14)
        self.lbl_prog = tk.Label(f4b, text="", font=("Calibri", 9),
                                 bg=self.LIGHT, fg=self.DARK)
        self.lbl_prog.pack(side="left")

        # ── 紀錄視窗 ──
        tk.Label(body, text="記錄", font=("Calibri", 9, "bold"),
                 bg=self.LIGHT, fg=self.DARK).pack(anchor="w", padx=8, pady=(8, 2))
        self.log = scrolledtext.ScrolledText(body, height=9,
                                             font=("Consolas", 8),
                                             bg="#F8F8F8", relief="flat",
                                             state="disabled")
        self.log.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        # 顏色標籤
        self.log.tag_config("ok",  foreground="#375623")
        self.log.tag_config("err", foreground=self.RED)
        self.log.tag_config("inf", foreground=self.DARK)

    def _step_frame(self, parent, title, idx):
        frm = tk.Frame(parent, bg=self.MED, pady=3, padx=8)
        frm.pack(fill="x", pady=(8 if idx > 0 else 0, 0))
        tk.Label(frm, text=title, font=("Calibri", 10, "bold"),
                 bg=self.MED, fg=self.WHITE).pack(anchor="w")

    def _step_body(self, parent, idx):
        frm = tk.Frame(parent, bg=self.LIGHT, pady=6, padx=8)
        frm.pack(fill="x")
        return frm

    def _btn_style(self, primary=False):
        return dict(
            font=("Calibri", 9, "bold"),
            bg=self.MED if primary else "#D9E2EC",
            fg=self.WHITE if primary else self.DARK,
            activebackground=self.DARK if primary else "#C0D0E0",
            relief="flat", padx=10, pady=3, cursor="hand2"
        )

    def _center_window(self):
        self.update_idletasks()
        w, h = 720, 620
        sw   = self.winfo_screenwidth()
        sh   = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    # ── 記錄輸出 ──────────────────────────────────────────

    def _log(self, msg, tag="inf"):
        self.log.config(state="normal")
        ts = datetime.now().strftime("%H:%M:%S")
        self.log.insert("end", f"[{ts}]  {msg}\n", tag)
        self.log.see("end")
        self.log.config(state="disabled")

    # ── 步驟 1：掃描 PDF ──────────────────────────────────

    def _browse_pdf(self):
        path = filedialog.askopenfilename(
            title="選擇 PDF 表格範本",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")])
        if path:
            self.pdf_path.set(path)

    def _scan_fields(self):
        path = self.pdf_path.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showwarning("提示", "請先選擇有效的 PDF 檔案。")
            return
        try:
            self._log(f"掃描中：{Path(path).name} …", "inf")
            self.fields = analyze_pdf(path)
            # 分類統計
            n_text = sum(1 for f in self.fields if f.field_type == "Text")
            n_cb   = sum(1 for f in self.fields if "CheckBox" in f.field_type)
            n_cmb  = sum(1 for f in self.fields if f.field_type == "Combo")
            summary = (f"✓ 偵測到 {len(self.fields)} 個欄位"
                       f"（文字 {n_text}  Checkbox {n_cb}  下拉 {n_cmb}）")
            self.lbl_scan.config(text=summary, fg=self.GREEN)
            self._log(summary, "ok")
        except Exception as e:
            self.lbl_scan.config(text=f"✗ 掃描失敗：{e}", fg=self.RED)
            self._log(f"掃描失敗：{e}", "err")

    # ── 步驟 2：匯出範本 ──────────────────────────────────

    def _export_template(self):
        if not self.fields:
            messagebox.showwarning("提示", "請先掃描 PDF 欄位（步驟①）。")
            return
        pdf_stem = Path(self.pdf_path.get()).stem
        out = filedialog.asksaveasfilename(
            title="儲存 Excel 範本",
            initialfile=f"{pdf_stem}_template.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not out:
            return
        try:
            generate_excel_template(self.fields, self.pdf_path.get(), out)
            self.lbl_export.config(text=f"✓ 已儲存：{Path(out).name}", fg=self.GREEN)
            self._log(f"Excel 範本已儲存：{out}", "ok")
            # 提示自動載入
            if messagebox.askyesno("完成", f"Excel 範本已生成。\n是否立即在 Excel 中開啟？"):
                os.startfile(out) if sys.platform == "win32" else os.system(f'open "{out}"')
        except Exception as e:
            self.lbl_export.config(text=f"✗ 生成失敗：{e}", fg=self.RED)
            self._log(f"生成失敗：{e}", "err")

    # ── 步驟 3：載入 Excel ────────────────────────────────

    def _browse_excel(self):
        path = filedialog.askopenfilename(
            title="選擇已填妥的 Excel",
            filetypes=[("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")])
        if path:
            self.excel_path.set(path)

    def _load_excel(self):
        path = self.excel_path.get().strip()
        if not path or not os.path.exists(path):
            messagebox.showwarning("提示", "請先選擇有效的 Excel 檔案。")
            return
        if not self.fields:
            messagebox.showwarning("提示", "請先掃描 PDF 欄位（步驟①）。")
            return
        try:
            _, self.rows = read_excel_data(path)
            non_empty = [r for r in self.rows
                         if any(_safe_str(v) for k, v in r.items()
                                if k != "__output_name__")]
            msg = f"✓ 讀取到 {len(non_empty)} 行有效資料"
            self.lbl_load.config(text=msg, fg=self.GREEN)
            self._log(msg, "ok")
            self.rows = non_empty
        except Exception as e:
            self.lbl_load.config(text=f"✗ 讀取失敗：{e}", fg=self.RED)
            self._log(f"讀取失敗：{e}", "err")

    # ── 步驟 4：生成 PDF ──────────────────────────────────

    def _browse_out(self):
        path = filedialog.askdirectory(title="選擇輸出資料夾")
        if path:
            self.out_dir.set(path)

    def _start_generate(self):
        # 驗證
        if not self.fields:
            messagebox.showwarning("提示", "請先完成步驟①掃描 PDF 欄位。")
            return
        if not self.rows:
            messagebox.showwarning("提示", "請先完成步驟③載入 Excel 資料。")
            return
        out = self.out_dir.get().strip()
        if not out:
            # 預設與 PDF 同目錄
            out = str(Path(self.pdf_path.get()).parent / "output_pdfs")
            self.out_dir.set(out)

        self.btn_gen.config(state="disabled")
        self.progress_var.set(0)
        self._log(f"開始生成 {len(self.rows)} 份 PDF → {out}", "inf")

        def _run():
            def _cb(done, total, fname, ok, err):
                pct = done / total * 100
                self.progress_var.set(pct)
                self.lbl_prog.config(text=f"{done}/{total}")
                tag = "ok" if ok else "err"
                msg = f"✓ {fname}" if ok else f"✗ {fname}  ({err})"
                self._log(msg, tag)

            results = batch_fill_pdfs(
                self.pdf_path.get(), self.fields,
                self.rows, out, progress_cb=_cb
            )
            ok_n  = sum(1 for _, ok, _ in results if ok)
            fail_n = len(results) - ok_n
            self._log(
                f"完成：{ok_n} 份成功  {fail_n} 份失敗  →  {out}",
                "ok" if fail_n == 0 else "err"
            )
            self.btn_gen.config(state="normal")
            self.progress_var.set(100)
            if messagebox.askyesno("完成", f"已生成 {ok_n}/{len(results)} 份 PDF。\n是否開啟輸出資料夾？"):
                if sys.platform == "win32":
                    os.startfile(out)
                else:
                    os.system(f'open "{out}"')

        threading.Thread(target=_run, daemon=True).start()


# ══════════════════════════════════════════════════════════
#  主程式
# ══════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = App()
    app.mainloop()
